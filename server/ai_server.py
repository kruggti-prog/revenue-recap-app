"""
FastAPI AI backend for Revenue Recap app.
Handles: image analysis (Claude vision), Excel parsing, audio transcription.
Runs on port 5001. The Express app proxies /api/ai/* to this server.
"""

import asyncio
import base64
import io
import os
import sys
from datetime import datetime
from typing import Optional

import anthropic
import openpyxl
import uvicorn
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse

# Add helpers to path
sys.path.insert(0, os.path.dirname(__file__))
from llm_helpers.transcribe_audio import transcribe_audio

app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

claude = anthropic.Anthropic()


def image_to_b64(data: bytes, media_type: str) -> str:
    return base64.b64encode(data).decode()


def analyze_with_vision(image_bytes: bytes, media_type: str, prompt: str) -> str:
    """Send image + prompt to Claude vision and return text response."""
    b64 = image_to_b64(image_bytes, media_type)
    msg = claude.messages.create(
        model="claude_sonnet_4_6",
        max_tokens=1024,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": b64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )
    return msg.content[0].text


# ─── Pickup & Pacing Analysis ────────────────────────────────────────────────

@app.post("/api/ai/pickup")
async def analyze_pickup(
    image: UploadFile = File(...),
):
    data = await image.read()
    media_type = image.content_type or "image/png"

    prompt = """Analyze this hotel Pickup & Pacing report screenshot.
Extract the data for the three months shown (Current Month, Next Month, Following Month).

For each month identify:
- Month name/label
- RN (Room Nights) pickup number
- Revenue pickup amount  
- ADR (Average Daily Rate)
- Revenue Delta pacing direction: look at the far-right column. If the value is positive/green/up, it's "up". If negative/red/down, it's "down". Include the actual value.

Return EXACTLY in this format (replace placeholders with real values):
Current Month: [month name] | RN: [number] | Rev: $[amount] | ADR: $[amount] | Delta pacing [up/down] [value]
Next Month: [month name] | RN: [number] | Rev: $[amount] | ADR: $[amount] | Delta pacing [up/down] [value]
Following Month: [month name] | RN: [number] | Rev: $[amount] | ADR: $[amount] | Delta pacing [up/down] [value]

If you cannot clearly read a value, use "N/A". Do not add any extra commentary."""

    try:
        result = analyze_with_vision(data, media_type, prompt)
        return {"analysis": result, "raw": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── CoStar Analysis ─────────────────────────────────────────────────────────

@app.post("/api/ai/costar")
async def analyze_costar(
    image: UploadFile = File(...),
    period: str = Form(...),  # "7" or "28"
):
    data = await image.read()
    media_type = image.content_type or "image/png"

    period_label = "last 7 days" if period == "7" else "last 28 days"

    prompt = f"""Analyze this CoStar hotel performance report screenshot. This data is from the {period_label}.

Write a brief summary for a hotel General Manager email. Focus on:
- Occupancy (Occ) performance vs comp set
- ADR performance vs comp set  
- RevPAR performance vs comp set

Format as bullet points. Keep it casual and positive in tone where the data supports it. 
If a metric is below comp set, frame it constructively (opportunity, focus area).
Be concise — 3-5 bullet points total. Start directly with the bullets, no intro sentence needed."""

    try:
        result = analyze_with_vision(data, media_type, prompt)
        return {"summary": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── OTA / Expedia Analysis ───────────────────────────────────────────────────

@app.post("/api/ai/ota")
async def analyze_ota(
    image: UploadFile = File(...),
    period: str = Form(...),
):
    data = await image.read()
    media_type = image.content_type or "image/png"

    period_labels = {
        "next14": "Stays Booked for the Next 14 Days vs Comp Set",
        "next28": "Stays Booked for the Next 28 Days vs Comp Set",
        "last7": "Stays from the Last 7 Days vs Comp Set",
        "last14": "Stays from the Last 14 Days vs Comp Set",
    }
    period_label = period_labels.get(period, period)

    prompt = f"""Analyze this Expedia OTA Production Report screenshot.
This is the "{period_label}" view.

Summarize the key findings vs the comp set. Include:
- How the property is performing vs competitors on Expedia
- Notable strengths or gaps
- Any visibility, pricing, or content opportunities you can identify

Keep it concise (3-5 bullet points), professional but conversational tone. 
Start directly with the bullets."""

    try:
        result = analyze_with_vision(data, media_type, prompt)
        return {"summary": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Business Mix Analysis ────────────────────────────────────────────────────

@app.post("/api/ai/business-mix")
async def analyze_business_mix(
    file: UploadFile = File(...),
):
    data = await file.read()
    today = datetime.now()
    day_of_month = today.day
    month_name = today.strftime("%B")
    date_str = today.strftime("%B %d, %Y")

    # Determine where we are in the month
    if day_of_month <= 5:
        timing_context = "We're just getting started — plenty of time to build pace and capitalize on demand."
    elif day_of_month <= 15:
        timing_context = "We're in the early-to-mid part of the month — still good time to influence the outcome."
    elif day_of_month <= 22:
        timing_context = "We're past the midpoint — still some opportunity to move the needle this month."
    else:
        timing_context = f"With only a few days left in {month_name}, this is more of a review — focus shifts to next month."

    # Try to parse Excel
    excel_text = ""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames[:3]:  # max 3 sheets
            ws = wb[sheet_name]
            excel_text += f"\n--- Sheet: {sheet_name} ---\n"
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_vals = [str(v) if v is not None else "" for v in row]
                if any(v.strip() for v in row_vals):
                    rows.append(" | ".join(row_vals))
            excel_text += "\n".join(rows[:80])  # cap at 80 rows per sheet
    except Exception as e:
        excel_text = f"[Could not parse file: {e}]"

    prompt = f"""You are analyzing a hotel Business Mix Report for a General Manager email recap.

Today's date: {date_str}
Timing note: {timing_context}

Here is the raw data from the Excel file:
{excel_text}

Instructions:
- Compare what's currently on the books to the same time last year (STLY)
- The columns labeled 'Final' show where last year ended — DO NOT include last year's final in the recap
- Summarize current pace vs STLY for key segments (Transient, Group, Contract, etc. — whatever is in the data)
- Note which segments are ahead, behind, or flat vs STLY
- Keep the tone casual and positive where applicable — this is for a GM
- Include the timing context naturally
- Format as 3-6 bullet points
- Start with: "As of {date_str}, here's a look at what's on the books:"

Keep it concise and conversational."""

    try:
        msg = claude.messages.create(
            model="claude_sonnet_4_6",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}],
        )
        summary = msg.content[0].text
        return {"summary": summary}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Audio / Transcript Transcription & Summary ───────────────────────────────

@app.post("/api/ai/transcribe")
async def transcribe_and_summarize(
    file: UploadFile = File(...),
    file_type: str = Form(default="audio"),  # "audio", "video", "transcript"
):
    data = await file.read()
    content_type = file.content_type or "audio/mpeg"

    if file_type == "transcript":
        # Plain text transcript — skip transcription
        transcript_text = data.decode("utf-8", errors="replace")
    else:
        # Audio or video — transcribe first
        # Map common video types to audio for transcription
        media_map = {
            "video/mp4": "audio/mp4",
            "video/webm": "audio/webm",
            "video/mpeg": "audio/mpeg",
            "audio/mp4": "audio/mp4",
            "audio/mpeg": "audio/mpeg",
            "audio/wav": "audio/wav",
            "audio/webm": "audio/webm",
            "audio/ogg": "audio/ogg",
            "audio/flac": "audio/flac",
        }
        media_type = media_map.get(content_type, "audio/mpeg")

        try:
            result = await transcribe_audio(data, media_type=media_type, diarize=True)
            transcript_text = result["text"]
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Transcription failed: {e}")

    # Now summarize the transcript as call notes
    prompt = f"""You are summarizing a hotel revenue management call for a General Manager email recap.

Here is the call transcript:
---
{transcript_text[:8000]}
---

Write a concise call summary in this format:
- Key topics discussed (bullet points)
- Action items or follow-ups (if mentioned)
- Any decisions made

Keep it professional but conversational. Start with "Here's a summary of today's call discussion:" 
Do not include filler or preamble."""

    try:
        msg = claude.messages.create(
            model="claude_sonnet_4_6",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}],
        )
        summary = msg.content[0].text
        return {"transcript": transcript_text, "summary": summary}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=5001)
